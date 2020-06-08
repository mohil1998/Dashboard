from django.shortcuts import render
import openpyxl
from .models import FilesDB,Dashboard,Graph
import datetime
import json
# Create your views here.

# file management
def connectdash(request):
    # when a list of files and dashboard are requested.
    if "GET" == request.method:
        user_files = FilesDB.objects.filter(user=request.user)
        user_dash = Dashboard.objects.filter(user=request.user)
        file_list = []
        dash_list = []
        for files in user_files:
            file_list.append(files.datafile)
        for dashes in user_dash:
            dash_list.append(dashes.dash_name)

        return render(request,'dash/connect_data.html',{"file_list":file_list,"dash_list":dash_list})

    else:
        excel_file = request.FILES["document"] # 'document' is the name of file when uploaded from frontend

        # you may put validations here to check extension or file size
        t=FilesDB.objects.get_or_create(user=request.user,datafile=excel_file) #to save the file in database
        user_files = FilesDB.objects.filter(user=request.user)
        file_list = []

        for files in user_files:
            file_list.append(files.datafile)

        #sending the updated list of file for the selection
        return render(request,'dash/connect_data.html',{"file_list":file_list})

#creating a new dashboard
def createdash(request):
    if "GET" == request.method:
        user_dash = Dashboard.objects.filter(user=request.user)
        dash_list = []
        for dashes in user_dash:
            dash_list.append(dashes.dash_name)
        return render(request,'dash/dash_home.html',{"dash_list":dash_list})
    else:
        name = request.POST['name'] # 'name' is name of name field name="name"
        description = request.POST['description'] # name="description"
        d = Dashboard.objects.get_or_create(dash_name=name,description=description,user=request.user) #request.user is used for getting current logged in user
        print("created")
        return render(request,'dash/dash_home.html',{})


def viewdash(request):
    if "GET" == request.method:
        user_dash = Dashboard.objects.filter(user=request.user)
        dash_list = []
        for dashes in user_dash:
            dash_list.append(dashes.dash_name)
        return render(request,'dash/view_dash.html',{"dash_list":dash_list})
    else:
        dash=request.POST['view_chart']
        d = Dashboard.objects.filter(dash_name=dash,user=request.user) # list of dashboards which satisfies the condition
        g = Graph.objects.filter(dash_name=d[0])
        print(g)
        index=1
        graphs={}
        objects=[]
        for idd in g:
            objects.append(idd)
            dict={}
            wb = openpyxl.load_workbook(idd.datafile)
            worksheet = wb.active
            print(type(worksheet))
            i=-1
            excel_data = list()
            for row in worksheet.iter_rows():
                for cell in row:
                    i=i+1
                    if str(cell.value) == idd.x_axis:
                        x=i
                    elif str(cell.value) == idd.y_axis:
                        y=i
                break

            m_row = worksheet.max_row # gives maximum number of rows
            m_col = worksheet.max_column # give maximum number of columns
            resx = isinstance(worksheet.cell(row = 2, column = x+1).value, datetime.datetime )
            resy = isinstance(worksheet.cell(row = 2, column = y+1).value, str)

            cell_obj = worksheet.cell(row = 3, column = x+1)
            #print(cell_obj.value.date.date)

            if str(resy)=='True':
                if str(resx)=='True':#check if xaxis is datefield
                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        dict[str(cell_obj.value.date())]=0
                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        cell_obj1 = worksheet.cell(row=j,column=y+1)
                        dict[str(cell_obj.value.date())]=dict[str(cell_obj.value.date())]+1
                else:
                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        dict[cell_obj.value]=0
                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        cell_obj1 = worksheet.cell(row=j,column=y+1)
                        dict[cell_obj.value]=dict[cell_obj.value]+1
            else:
                #initializing all the key
                if str(resx)=='True':#check if xaxis is datefield
                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        dict[str(cell_obj.value.date())]=0

                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        cell_obj1 = worksheet.cell(row=j,column=y+1)
                        dict[str(cell_obj.value.date())]=dict[str(cell_obj.value.date())]+cell_obj1.value
                else:
                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        dict[cell_obj.value]=0

                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        cell_obj1 = worksheet.cell(row=j,column=y+1)
                        dict[cell_obj.value]=dict[cell_obj.value]+cell_obj1.value
            print(dict)
            graphs[index]={'type':idd.graph_type,'dict':dict}
            print(graphs)
            index=index+1
        return render(request,'dash/view_dash.html',{'data':graphs,'dash_name':dash})
