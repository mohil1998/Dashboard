import asyncio
import numpy as np
import random
import json
from channels.consumer import AsyncConsumer
from .models import FilesDB,Dashboard,Graph
import openpyxl
import matplotlib.pyplot as plt
import datetime

class ChatConsumer(AsyncConsumer):

    async def websocket_connect(self, event):
        print("connected", event)
        await self.send({
            "type": "websocket.accept"
        })

    async def websocket_receive(self, event):
        print("receive", event)
        data = json.loads(event['text'])

        print(data)
        if data['choice']=='1':  #choice 1 is for sending the excel data to frontend
            wb = openpyxl.load_workbook(data['file'])
            worksheet = wb.active
            print(worksheet)
            excel_data = list()
            #
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)

            dict = {'excel_data':excel_data,'choice':'1'}
            await self.send({
                'type': 'websocket.send',
                'text': json.dumps(dict),
            })

        elif data['choice']=='2': # for creating graphs
            x,y=0,0
            dict={}
            year=[]
            month=[]

            d = Dashboard.objects.filter(dash_name=data['dash'],user=self.scope["user"])
            print(d[0])
            g=Graph.objects.get_or_create(dash_name=d[0],
                                        datafile=data['file'],
                                        graph_type=data['graph'],
                                        x_axis=data['xaxis'],
                                        y_axis=data['yaxis'])
            print("graph entry done")

            wb = openpyxl.load_workbook(data['file'])
            worksheet = wb.active
            print(type(worksheet))
            i=-1
            excel_data = list()
            for row in worksheet.iter_rows():
                for cell in row:
                    i=i+1
                    if str(cell.value) == data['xaxis']:
                        x=i
                    elif str(cell.value) == data['yaxis']:
                        y=i
                break

            m_row = worksheet.max_row
            m_col = worksheet.max_column
            print(m_col,m_row,x,y)
            print(str(type(worksheet.cell(row = 2, column = x+1).value)))
            resx = isinstance(worksheet.cell(row = 2, column = x+1).value, datetime.datetime )
            resy = isinstance(worksheet.cell(row = 2, column = y+1).value, str)


            cell_obj = worksheet.cell(row = 3, column = x+1)
            #print(cell_obj.value.date.date)

            if str(resy)=='True':
                if str(resx)=='True':#check if xaxis is datefield
                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        dict[str(cell_obj.value.date())]=0
                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        cell_obj1 = worksheet.cell(row=j,column=y+1)
                        dict[str(cell_obj.value.date())]=dict[str(cell_obj.value.date())]+1
                else:
                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        dict[cell_obj.value]=0

                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        cell_obj1 = worksheet.cell(row=j,column=y+1)
                        dict[cell_obj.value]=dict[cell_obj.value]+1
            else:
                #initializing all the key
                if str(resx)=='True':#check if xaxis is datefield
                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        dict[str(cell_obj.value.date())]=0

                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        cell_obj1 = worksheet.cell(row=j,column=y+1)
                        dict[str(cell_obj.value.date())]=dict[str(cell_obj.value.date())]+cell_obj1.value
                else:
                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        dict[cell_obj.value]=0

                    for j in range(2,m_row+1):
                        cell_obj = worksheet.cell(row = j, column = x+1)
                        cell_obj1 = worksheet.cell(row=j,column=y+1)
                        if cell_obj1==None:
                            cell_obj1=0
                        dict[cell_obj.value]=dict[cell_obj.value]+cell_obj1.value


            dict['choice']='2'
            dict['xaxis']=data['xaxis']
            dict['yaxis']=data['yaxis']

            await self.send({
                'type': 'websocket.send',
                'text': json.dumps(dict),
            })

        elif data['choice']=='3': # choice is for deleting a graph from database
            # print(data['dict'])
            d = Dashboard.objects.filter(dash_name=data['dict']['dash'],user=self.scope["user"])
            print(d[0])
            instance = Graph.objects.filter(dash_name=d[0],
                                            datafile=data['dict']['file'],
                                            graph_type=data['dict']['graph'],
                                            x_axis=data['dict']['xaxis'],
                                            y_axis=data['dict']['yaxis'])
            instance[0].delete()
            print("deleted")

        elif data['choice']=='4':
            d = Dashboard.objects.filter(dash_name=data['dash'],user=self.scope["user"]) # list of dashboards which satisfies the condition
            g = Graph.objects.filter(dash_name=d[0])
            # print(g)
            index=1
            graphs={}
            objects=[]
            for idd in g:
                objects.append(idd)
                dict={}
                wb = openpyxl.load_workbook(idd.datafile)
                worksheet = wb.active
                # print(type(worksheet))
                i=-1
                excel_data = list()
                for row in worksheet.iter_rows():
                    for cell in row:
                        i=i+1
                        if str(cell.value) == idd.x_axis:
                            x=i
                        elif str(cell.value) == idd.y_axis:
                            y=i
                    break

                m_row = worksheet.max_row # gives maximum number of rows
                m_col = worksheet.max_column # give maximum number of columns
                resx = isinstance(worksheet.cell(row = 2, column = x+1).value, datetime.datetime )
                resy = isinstance(worksheet.cell(row = 2, column = y+1).value, str)

                cell_obj = worksheet.cell(row = 3, column = x+1)
                if str(resy)=='True':
                    if str(resx)=='True':#check if xaxis is datefield
                        for j in range(2,m_row+1):
                            cell_obj = worksheet.cell(row = j, column = x+1)
                            dict[str(cell_obj.value.date())]=0
                        for j in range(2,m_row+1):
                            cell_obj = worksheet.cell(row = j, column = x+1)
                            cell_obj1 = worksheet.cell(row=j,column=y+1)
                            dict[str(cell_obj.value.date())]=dict[str(cell_obj.value.date())]+1
                    else:
                        for j in range(2,m_row+1):
                            cell_obj = worksheet.cell(row = j, column = x+1)
                            dict[cell_obj.value]=0
                        for j in range(2,m_row+1):
                            cell_obj = worksheet.cell(row = j, column = x+1)
                            cell_obj1 = worksheet.cell(row=j,column=y+1)
                            dict[cell_obj.value]=dict[cell_obj.value]+1
                else:
                    #initializing all the key
                    if str(resx)=='True':#check if xaxis is datefield
                        for j in range(2,m_row+1):
                            cell_obj = worksheet.cell(row = j, column = x+1)
                            dict[str(cell_obj.value.date())]=0

                        for j in range(2,m_row+1):
                            cell_obj = worksheet.cell(row = j, column = x+1)
                            cell_obj1 = worksheet.cell(row=j,column=y+1)
                            dict[str(cell_obj.value.date())]=dict[str(cell_obj.value.date())]+cell_obj1.value
                    else:
                        for j in range(2,m_row+1):
                            cell_obj = worksheet.cell(row = j, column = x+1)
                            if cell_obj=='None':
                                cell_obj.value=0
                            dict[cell_obj.value]=0

                        for j in range(2,m_row+1):
                            cell_obj = worksheet.cell(row = j, column = x+1)
                            cell_obj1 = worksheet.cell(row=j,column=y+1)
                            # print(dict[cell_obj.value],cell_obj1.value)
                            if cell_obj1=='None':
                                cell_obj1.value=0
                            if cell_obj=='None':
                                cell_obj.value=0
                            print('----------------------------------')
                            print(cell_obj.value,cell_obj1.value)
                            dict[cell_obj.value]=dict[cell_obj.value]+cell_obj1.value

                graphs[index]={'type':idd.graph_type,'dict':dict}
                index=index+1


            await self.send({
                'type': 'websocket.send',
                'text': json.dumps({'data':graphs,'choice':'4'}),
            })

    async def websocket_disconnect(self, event):
        print("disconnected", event)


class dashConsumer(AsyncConsumer):
    async def websocket_connect(self, event):
        print("connected", event)

        await self.send({
            "type": "websocket.accept"
        })

    async def websocket_receive(self, event):
        # print("receive", event)
        data = json.loads(event['text'])
        # print(data)
        d = Dashboard.objects.filter(dash_name=data['dash_name'],user=self.scope['user'])
        g = Graph.objects.filter(dash_name=d[0])
        print(g[int(data['id'])])
        g[int(data['id'])].delete()
        print("deleted")

    async def websocket_disconnect(self, event):
        print("disconnected", event)
